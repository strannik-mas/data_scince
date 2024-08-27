import csv
import datetime as dt
import openpyxl

# name.csv ниже нужно заменить на название csv файла, который вы хотите открыть
file_path = 'click_stream.csv'
with open(file_path) as csv_file:  # открываем файл
    # csv_reader = csv.DictReader(csv_file, fieldnames=['ID', 'page', 'date', 'device','sex'], dialect='excel')  # читаем файл
    data = csv.reader(csv_file, delimiter=';')
    months = ('jan', 'feb', 'mar', 'apr')
    pages = ('1_home_page', '2_search_page', '3_payment_page', '4_payment_confirmation_page')

    funnel = {}

    # for row in csv_reader:  # перебираем по одной строчке нашего файла
    for row in data:
        # page = list(row.items())[1][1]  # так можно получить первый (не нулевой) элемент строки -
        # для нашего датасета это строка, указывающая, на какой страничке было совершено действие
        # print(page)
        # ваш код для расчета воронки
        date = dt.date.fromisoformat(row[2])
        if row[4] not in funnel.keys():
            funnel[row[4]] = {
                row[3]: {
                    months[date.month - 1]: {
                        row[1]: row
                    }
                }
            }
        elif row[3] not in funnel[row[4]].keys():
            funnel[row[4]][row[3]] = {
                    months[date.month - 1]: {
                        row[1]: row
                    }
                }
        elif months[date.month - 1] not in funnel[row[4]][row[3]].keys():
            funnel[row[4]][row[3]][months[date.month - 1]] = {
                        row[1]: row
                    }
        elif row[1] not in funnel[row[4]][row[3]][months[date.month - 1]].keys():
            funnel[row[4]][row[3]][months[date.month - 1]][row[1]] = row
        else:
            funnel[row[4]][row[3]][months[date.month - 1]][row[1]].append(row)

wb = openpyxl.load_workbook("lesson_7.xlsx")
# print(funnel)
for sex, sexDict in funnel.items():
    stepRows = 1
    if sex not in wb.worksheets:
        wb.create_sheet(sex)
    ws = wb[sex]
    # print(f'Для пола {sex}')
    for device, deviceDict in sexDict.items():
        ws.cell(stepRows,2, device)
        ws.cell(2 + stepRows, 1, pages[0])
        ws.cell(3 + stepRows, 1, pages[1])
        ws.cell(4 + stepRows, 1, pages[2])
        ws.cell(5 + stepRows, 1, pages[3])
        ws.cell(6 + stepRows, 1, 'CONVERSION PERCENTAGE')
        ws.merge_cells(f'B{stepRows}:E{stepRows}')
        # print(f'Для девайса {device}')
        for month, lists in deviceDict.items():
            colNum = months.index(month) + 2
            ws.cell(1 + stepRows, colNum, month)
            ws.cell(2 + stepRows, colNum, str(len(lists[pages[0]])))
            ws.cell(3 + stepRows, colNum, str(len(lists[pages[1]])))
            ws.cell(4 + stepRows, colNum, str(len(lists[pages[2]])))
            ws.cell(5 + stepRows, colNum, str(len(lists[pages[3]])))
            ws.cell(6 + stepRows, colNum, str(round(len(lists[pages[3]]) * 100 / len(lists[pages[0]]), 2)).replace('.', ','))
        # print(f'Main page count in {month}: {len(lists[pages[0]])}')
        # print(f'Search page count in {month}: {len(lists[pages[1]])}')
        # print(f'Payment page count in {month}: {len(lists[pages[2]])}')
        # print(f'Payment confirmation page count in {month}: {len(lists[pages[3]])}')
        # print(f'\nПРОЦЕНТ КОНВЕРСИИ in {month}: {round(len(lists[pages[3]]) * 100 / len(lists[pages[0]]), 2)}%\n\n')
        stepRows += 9
    wb.save("lesson_7.xlsx")

